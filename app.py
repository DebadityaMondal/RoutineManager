"""
School Routine Generator - Flask Web Application with MySQL & User Auth
"""

import json
import io
from flask import Flask, render_template, request, redirect, url_for, flash, session
from flask_login import LoginManager, login_user, logout_user, login_required, current_user

from config import Config
from models import (
    db, User, Subject, Teacher, TeacherSubject,
    ClassSection, ScheduleSettings, GeneratedRoutine, ClassSubjectPeriod, ClassTeacher,
)
from generator import RoutineGenerator, SchoolInputData, TeacherData

app = Flask(__name__)
app.config.from_object(Config)

db.init_app(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"
login_manager.login_message_category = "warning"


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


def get_or_create_settings(user_id):
    """Get or create schedule settings for a user."""
    settings = ScheduleSettings.query.filter_by(user_id=user_id).first()
    if not settings:
        settings = ScheduleSettings(user_id=user_id)
        db.session.add(settings)
        db.session.commit()
    return settings


# ─── Auth Routes ──────────────────────────────────────────────

@app.route("/register", methods=["GET", "POST"])
def register():
    if current_user.is_authenticated:
        return redirect(url_for("index"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "").strip()
        confirm = request.form.get("confirm_password", "").strip()

        if not username or not email or not password:
            flash("All fields are required.", "danger")
            return redirect(url_for("register"))

        if password != confirm:
            flash("Passwords do not match.", "danger")
            return redirect(url_for("register"))

        if User.query.filter_by(username=username).first():
            flash("Username already taken.", "danger")
            return redirect(url_for("register"))

        if User.query.filter_by(email=email).first():
            flash("Email already registered.", "danger")
            return redirect(url_for("register"))

        user = User(username=username, email=email)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()

        flash("Registration successful! Please log in.", "success")
        return redirect(url_for("login"))

    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user)
            flash(f"Welcome back, {user.username}!", "success")
            next_page = request.args.get("next")
            return redirect(next_page or url_for("index"))
        else:
            flash("Invalid username or password.", "danger")

    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("Logged out successfully.", "success")
    return redirect(url_for("login"))


# ─── Main Routes ─────────────────────────────────────────────

@app.route("/")
@login_required
def index():
    subjects = Subject.query.filter_by(user_id=current_user.id).all()
    teachers = Teacher.query.filter_by(user_id=current_user.id).all()
    classes = ClassSection.query.filter_by(user_id=current_user.id).all()
    settings = get_or_create_settings(current_user.id)
    gen_routine = GeneratedRoutine.query.filter_by(user_id=current_user.id).first()

    routine = gen_routine.routine_data if gen_routine else None
    # Build day info
    day_info = {}
    max_periods = 0
    for day in settings.days_list:
        num = settings.get_periods_for_day(day)
        day_info[day] = {"periods": num, "has_tiffin": settings.has_tiffin(day)}
        if num > max_periods:
            max_periods = num

    # Substitute data from session
    sub_result = session.pop("substitute_result", None)
    sub_teacher = session.pop("substitute_teacher", "")
    sub_day = session.pop("substitute_day", "")
    teacher_names = [t.name for t in teachers]
    tiffin_after = settings.tiffin_after_period or 0

    return render_template("index.html", subjects=subjects, teachers=teachers,
                           classes=classes, settings=settings, routine=routine,
                           day_info=day_info, max_periods=max_periods,
                           teacher_names=teacher_names, tiffin_after=tiffin_after,
                           sub_result=sub_result, sub_teacher=sub_teacher, sub_day=sub_day)


@app.route("/add_subject", methods=["POST"])
@login_required
def add_subject():
    name = request.form.get("subject", "").strip()
    if name:
        existing = Subject.query.filter_by(name=name, user_id=current_user.id).first()
        if existing:
            flash(f"Subject '{name}' already exists.", "warning")
        else:
            db.session.add(Subject(name=name, user_id=current_user.id))
            db.session.commit()
            flash(f"Subject '{name}' added.", "success")
    return redirect(url_for("index"))


@app.route("/delete_subject/<int:subject_id>")
@login_required
def delete_subject(subject_id):
    subject = Subject.query.filter_by(id=subject_id, user_id=current_user.id).first()
    if subject:
        db.session.delete(subject)
        db.session.commit()
        flash(f"Subject '{subject.name}' removed.", "success")
    return redirect(url_for("index"))


@app.route("/add_teacher", methods=["POST"])
@login_required
def add_teacher():
    name = request.form.get("teacher_name", "").strip()
    short_name = request.form.get("teacher_short_name", "").strip()
    subjects = request.form.getlist("teacher_subjects")
    if name and subjects:
        existing = Teacher.query.filter_by(name=name, user_id=current_user.id).first()
        if existing:
            flash(f"Teacher '{name}' already exists.", "warning")
        else:
            teacher = Teacher(name=name, short_name=short_name or None, user_id=current_user.id)
            db.session.add(teacher)
            db.session.flush()
            for s in subjects:
                db.session.add(TeacherSubject(teacher_id=teacher.id, subject_name=s))
            db.session.commit()
            flash(f"Teacher '{name}' added.", "success")
    else:
        flash("Please provide teacher name and select at least one subject.", "danger")
    return redirect(url_for("index"))


@app.route("/delete_teacher/<int:teacher_id>")
@login_required
def delete_teacher(teacher_id):
    teacher = Teacher.query.filter_by(id=teacher_id, user_id=current_user.id).first()
    if teacher:
        db.session.delete(teacher)
        db.session.commit()
        flash(f"Teacher '{teacher.name}' removed.", "success")
    return redirect(url_for("index"))


@app.route("/add_class", methods=["POST"])
@login_required
def add_class():
    class_name = request.form.get("class_name", "").strip()
    sections = request.form.get("sections", "").strip()
    if class_name and sections:
        for section in sections.split(","):
            section = section.strip()
            if section:
                existing = ClassSection.query.filter_by(
                    class_name=class_name, section=section, user_id=current_user.id
                ).first()
                if not existing:
                    db.session.add(ClassSection(
                        class_name=class_name, section=section, user_id=current_user.id
                    ))
        db.session.commit()
        flash(f"Class '{class_name}' added.", "success")
    else:
        flash("Please provide class name and sections.", "danger")
    return redirect(url_for("index"))


@app.route("/delete_class/<int:class_id>")
@login_required
def delete_class(class_id):
    cls = ClassSection.query.filter_by(id=class_id, user_id=current_user.id).first()
    if cls:
        db.session.delete(cls)
        db.session.commit()
        flash(f"'{cls.full_name}' removed.", "success")
    return redirect(url_for("index"))


@app.route("/settings", methods=["GET", "POST"])
@login_required
def update_settings():
    settings = get_or_create_settings(current_user.id)
    subjects = Subject.query.filter_by(user_id=current_user.id).all()

    if request.method == "POST":
        periods = request.form.get("periods_per_day", "6").strip()
        try:
            settings.periods_per_day = int(periods)
        except ValueError:
            settings.periods_per_day = 6

        days = request.form.get("days", "").strip()
        if days:
            settings.days_list = [d.strip() for d in days.split(",") if d.strip()]

        # Day-wise periods
        day_periods = {}
        for day in settings.days_list:
            key = f"periods_{day}"
            val = request.form.get(key, "").strip()
            if val:
                try:
                    day_periods[day] = int(val)
                except ValueError:
                    pass
        settings.day_periods = day_periods

        # First period preferences
        first_period = request.form.getlist("first_period_subjects")
        settings.first_period_subjects = first_period

        # Class teacher toggle
        settings.enable_class_teacher = "enable_class_teacher" in request.form

        # Teacher workload soft constraints
        max_tp = request.form.get("max_teacher_periods_per_day", "0").strip()
        try:
            settings.max_teacher_periods_per_day = int(max_tp) if max_tp else 0
        except ValueError:
            settings.max_teacher_periods_per_day = 0
        settings.avoid_consecutive = "avoid_consecutive" in request.form

        # Tiffin break setting
        tiffin_val = request.form.get("tiffin_after_period", "0").strip()
        try:
            settings.tiffin_after_period = int(tiffin_val) if tiffin_val else 0
        except ValueError:
            settings.tiffin_after_period = 0

        db.session.commit()
        flash("Settings updated.", "success")
        return redirect(url_for("update_settings"))

    classes = ClassSection.query.filter_by(user_id=current_user.id).all()
    teachers = Teacher.query.filter_by(user_id=current_user.id).all()
    ct_records = ClassTeacher.query.filter_by(user_id=current_user.id).all()
    ct_map = {rec.class_name: {"teacher": rec.teacher_name, "subject": rec.subject_name} for rec in ct_records}

    # Build teacher->subjects JSON for dependent dropdown
    teacher_subjects_json = json.dumps({t.name: t.subject_names for t in teachers})
    ct_map_json = json.dumps(ct_map)

    return render_template("settings.html", settings=settings, subjects=subjects,
                           classes=classes, teachers=teachers, ct_map=ct_map,
                           teacher_subjects_json=teacher_subjects_json,
                           ct_map_json=ct_map_json)


# ─── Class Teacher Assignment ─────────────────────────────────

@app.route("/save_class_teachers", methods=["POST"])
@login_required
def save_class_teachers():
    """Save class teacher assignments with validation."""
    classes = ClassSection.query.filter_by(user_id=current_user.id).all()
    teachers = Teacher.query.filter_by(user_id=current_user.id).all()
    teacher_subject_map = {t.name: set(t.subject_names) for t in teachers}

    # Collect assignments and validate
    assignments = []
    teacher_usage = {}  # teacher_name -> class_name (for duplicate check)
    errors = []

    for cls in classes:
        teacher_key = f"ct_teacher_{cls.full_name}"
        subject_key = f"ct_subject_{cls.full_name}"
        teacher_name = request.form.get(teacher_key, "").strip()
        subject_name = request.form.get(subject_key, "").strip()

        if teacher_name and subject_name:
            # Check if this teacher is already assigned to another class
            if teacher_name in teacher_usage:
                errors.append(
                    f"'{teacher_name}' is assigned to both '{teacher_usage[teacher_name]}' "
                    f"and '{cls.full_name}'. A teacher can only be class teacher of one class."
                )
            else:
                teacher_usage[teacher_name] = cls.full_name

            # Validate subject belongs to teacher
            if teacher_name in teacher_subject_map:
                if subject_name not in teacher_subject_map[teacher_name]:
                    errors.append(
                        f"'{cls.full_name}': '{teacher_name}' does not teach '{subject_name}'. "
                        f"Available: {', '.join(teacher_subject_map[teacher_name])}."
                    )
            else:
                errors.append(f"'{cls.full_name}': Teacher '{teacher_name}' not found.")

            if not errors or teacher_name not in teacher_usage:
                assignments.append({
                    "class_name": cls.full_name,
                    "teacher_name": teacher_name,
                    "subject_name": subject_name,
                })
        elif teacher_name and not subject_name:
            errors.append(f"'{cls.full_name}': Please select a subject for '{teacher_name}'.")
        elif subject_name and not teacher_name:
            errors.append(f"'{cls.full_name}': Please select a teacher for subject '{subject_name}'.")

    if errors:
        for err in errors:
            flash(err, "danger")
        return redirect(url_for("update_settings"))

    # All valid — save
    ClassTeacher.query.filter_by(user_id=current_user.id).delete()
    for a in assignments:
        db.session.add(ClassTeacher(
            user_id=current_user.id,
            class_name=a["class_name"],
            teacher_name=a["teacher_name"],
            subject_name=a["subject_name"],
        ))

    db.session.commit()
    flash("Class teacher assignments saved.", "success")
    return redirect(url_for("update_settings"))


# ─── Class-wise Subject Period Config ─────────────────────────

@app.route("/class_subject_periods", methods=["GET", "POST"])
@login_required
def class_subject_periods():
    """Configure how many periods per week each class gets for each subject."""
    subjects = Subject.query.filter_by(user_id=current_user.id).all()
    classes = ClassSection.query.filter_by(user_id=current_user.id).all()
    existing = ClassSubjectPeriod.query.filter_by(user_id=current_user.id).all()

    # Build lookup: {class_full_name: {subject_name: periods}}
    csp_map = {}
    for rec in existing:
        csp_map.setdefault(rec.class_name, {})[rec.subject_name] = rec.periods_per_week

    if request.method == "POST":
        # Clear existing
        ClassSubjectPeriod.query.filter_by(user_id=current_user.id).delete()

        # Save new values
        for cls in classes:
            for subj in subjects:
                key = f"csp_{cls.full_name}_{subj.name}"
                val = request.form.get(key, "").strip()
                if val:
                    try:
                        periods = int(val)
                        # -1 means not applicable, positive means fixed periods
                        if periods == -1 or periods > 0:
                            db.session.add(ClassSubjectPeriod(
                                user_id=current_user.id,
                                class_name=cls.full_name,
                                subject_name=subj.name,
                                periods_per_week=periods,
                            ))
                    except ValueError:
                        pass

        db.session.commit()
        flash("Class-wise subject periods updated.", "success")
        return redirect(url_for("class_subject_periods"))

    return render_template("class_subject_periods.html",
                           subjects=subjects, classes=classes, csp_map=csp_map)


# ─── Profile Routes ──────────────────────────────────────────

@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    """Edit user profile (username, email)."""
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip()

        if not username or not email:
            flash("Username and email are required.", "danger")
            return redirect(url_for("profile"))

        # Check uniqueness (exclude current user)
        if username != current_user.username:
            if User.query.filter_by(username=username).first():
                flash("Username already taken.", "danger")
                return redirect(url_for("profile"))

        if email != current_user.email:
            if User.query.filter_by(email=email).first():
                flash("Email already registered.", "danger")
                return redirect(url_for("profile"))

        current_user.username = username
        current_user.email = email
        db.session.commit()
        flash("Profile updated successfully.", "success")
        return redirect(url_for("profile"))

    return render_template("profile.html")


@app.route("/change_password", methods=["POST"])
@login_required
def change_password():
    """Change user password."""
    current_password = request.form.get("current_password", "").strip()
    new_password = request.form.get("new_password", "").strip()
    confirm_password = request.form.get("confirm_password", "").strip()

    if not current_password or not new_password:
        flash("All password fields are required.", "danger")
        return redirect(url_for("profile"))

    if not current_user.check_password(current_password):
        flash("Current password is incorrect.", "danger")
        return redirect(url_for("profile"))

    if new_password != confirm_password:
        flash("New passwords do not match.", "danger")
        return redirect(url_for("profile"))

    if len(new_password) < 4:
        flash("Password must be at least 4 characters.", "danger")
        return redirect(url_for("profile"))

    current_user.set_password(new_password)
    db.session.commit()
    flash("Password changed successfully.", "success")
    return redirect(url_for("profile"))


# ─── Edit Subject & Teacher Routes ───────────────────────────

@app.route("/edit_subject/<int:subject_id>", methods=["GET", "POST"])
@login_required
def edit_subject(subject_id):
    """Edit a subject name."""
    subject = Subject.query.filter_by(id=subject_id, user_id=current_user.id).first_or_404()

    if request.method == "POST":
        new_name = request.form.get("name", "").strip()
        if not new_name:
            flash("Subject name cannot be empty.", "danger")
            return redirect(url_for("edit_subject", subject_id=subject_id))

        # Check for duplicate
        existing = Subject.query.filter_by(name=new_name, user_id=current_user.id).first()
        if existing and existing.id != subject_id:
            flash(f"Subject '{new_name}' already exists.", "warning")
            return redirect(url_for("edit_subject", subject_id=subject_id))

        old_name = subject.name
        subject.name = new_name

        # Update teacher_subjects references
        TeacherSubject.query.filter(
            TeacherSubject.teacher_id.in_(
                db.session.query(Teacher.id).filter_by(user_id=current_user.id)
            ),
            TeacherSubject.subject_name == old_name
        ).update({TeacherSubject.subject_name: new_name}, synchronize_session="fetch")

        db.session.commit()
        flash(f"Subject renamed from '{old_name}' to '{new_name}'.", "success")
        return redirect(url_for("index"))

    return render_template("edit_subject.html", subject=subject)


@app.route("/edit_teacher/<int:teacher_id>", methods=["GET", "POST"])
@login_required
def edit_teacher(teacher_id):
    """Edit a teacher's name and subjects."""
    teacher = Teacher.query.filter_by(id=teacher_id, user_id=current_user.id).first_or_404()
    subjects = Subject.query.filter_by(user_id=current_user.id).all()

    if request.method == "POST":
        new_name = request.form.get("teacher_name", "").strip()
        new_subjects = request.form.getlist("teacher_subjects")

        if not new_name:
            flash("Teacher name cannot be empty.", "danger")
            return redirect(url_for("edit_teacher", teacher_id=teacher_id))

        # Check for duplicate name
        existing = Teacher.query.filter_by(name=new_name, user_id=current_user.id).first()
        if existing and existing.id != teacher_id:
            flash(f"Teacher '{new_name}' already exists.", "warning")
            return redirect(url_for("edit_teacher", teacher_id=teacher_id))

        if not new_subjects:
            flash("Please select at least one subject.", "danger")
            return redirect(url_for("edit_teacher", teacher_id=teacher_id))

        teacher.name = new_name
        teacher.short_name = request.form.get("teacher_short_name", "").strip() or None

        # Replace teacher subjects
        TeacherSubject.query.filter_by(teacher_id=teacher.id).delete()
        for s in new_subjects:
            db.session.add(TeacherSubject(teacher_id=teacher.id, subject_name=s))

        db.session.commit()
        flash(f"Teacher '{new_name}' updated.", "success")
        return redirect(url_for("index"))

    return render_template("edit_teacher.html", teacher=teacher, subjects=subjects)


@app.route("/upload", methods=["POST"])
@login_required
def upload_json():
    """Upload a JSON file to populate all school data."""
    file = request.files.get("json_file")
    if not file or not file.filename.endswith(".json"):
        flash("Please upload a valid .json file.", "danger")
        return redirect(url_for("index"))

    try:
        raw = json.load(file)

        # Clear existing data for this user
        Subject.query.filter_by(user_id=current_user.id).delete()
        Teacher.query.filter_by(user_id=current_user.id).delete()
        ClassSection.query.filter_by(user_id=current_user.id).delete()

        # Subjects
        for s in raw.get("subjects", []):
            db.session.add(Subject(name=s, user_id=current_user.id))

        # Teachers
        for t in raw.get("teachers", []):
            teacher = Teacher(name=t["name"], user_id=current_user.id)
            db.session.add(teacher)
            db.session.flush()
            for s in t.get("subjects", []):
                db.session.add(TeacherSubject(teacher_id=teacher.id, subject_name=s))

        # Classes
        for c in raw.get("classes", []):
            for section in c.get("sections", ["A"]):
                db.session.add(ClassSection(
                    class_name=c["name"], section=section, user_id=current_user.id
                ))

        # Settings
        settings = get_or_create_settings(current_user.id)
        settings.periods_per_day = raw.get("periods_per_day", 6)
        days = raw.get("days", ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"])
        settings.days_list = days
        settings.day_periods = raw.get("day_periods", {})
        settings.first_period_subjects = raw.get("first_period_subjects", [])

        db.session.commit()
        flash(f"Data loaded from '{file.filename}' successfully!", "success")
    except (json.JSONDecodeError, KeyError) as e:
        db.session.rollback()
        flash(f"Error parsing JSON: {e}", "danger")

    return redirect(url_for("index"))


@app.route("/upload_excel", methods=["POST"])
@login_required
def upload_excel():
    """Upload Excel/CSV file with teacher-subject and class-section sheets."""
    import openpyxl

    file = request.files.get("excel_file")
    if not file:
        flash("Please select a file.", "danger")
        return redirect(url_for("index"))

    filename = file.filename.lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xls") or filename.endswith(".csv")):
        flash("Please upload a valid .xlsx or .csv file.", "danger")
        return redirect(url_for("index"))

    try:
        if filename.endswith(".csv"):
            # Parse CSV as teacher-subject mapping
            import csv
            content = file.read().decode("utf-8")
            reader = csv.DictReader(io.StringIO(content))
            _process_teacher_subject_csv(reader)
        else:
            # Parse Excel workbook
            wb = openpyxl.load_workbook(file, data_only=True)
            sheet_names = [s.lower() for s in wb.sheetnames]

            # Look for teacher-subject sheet
            ts_sheet = None
            for name in wb.sheetnames:
                if "teacher" in name.lower() or "subject" in name.lower():
                    ts_sheet = wb[name]
                    break
            if not ts_sheet and len(wb.sheetnames) >= 1:
                ts_sheet = wb[wb.sheetnames[0]]

            # Look for class-section sheet
            cs_sheet = None
            for name in wb.sheetnames:
                if "class" in name.lower() or "section" in name.lower():
                    cs_sheet = wb[name]
                    break
            if not cs_sheet and len(wb.sheetnames) >= 2:
                cs_sheet = wb[wb.sheetnames[1]]

            if ts_sheet:
                _process_teacher_subject_sheet(ts_sheet)
            if cs_sheet:
                _process_class_section_sheet(cs_sheet)

        db.session.commit()
        flash(f"Data loaded from '{file.filename}' successfully!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error processing file: {e}", "danger")

    return redirect(url_for("index"))


def _process_teacher_subject_csv(reader):
    """Process CSV rows with columns: Teacher, Subjects (comma-separated)."""
    subjects_added = set()
    for row in reader:
        teacher_name = (row.get("Teacher") or row.get("teacher") or "").strip()
        subjects_str = (row.get("Subjects") or row.get("subjects") or "").strip()
        if not teacher_name or not subjects_str:
            continue

        subject_list = [s.strip() for s in subjects_str.split(",") if s.strip()]

        # Add subjects if new
        for s in subject_list:
            if s not in subjects_added:
                existing = Subject.query.filter_by(name=s, user_id=current_user.id).first()
                if not existing:
                    db.session.add(Subject(name=s, user_id=current_user.id))
                subjects_added.add(s)

        # Add teacher
        existing_t = Teacher.query.filter_by(name=teacher_name, user_id=current_user.id).first()
        if not existing_t:
            teacher = Teacher(name=teacher_name, user_id=current_user.id)
            db.session.add(teacher)
            db.session.flush()
            for s in subject_list:
                db.session.add(TeacherSubject(teacher_id=teacher.id, subject_name=s))


def _process_teacher_subject_sheet(sheet):
    """Process Excel sheet: Column A = Teacher Name, Column B = Subjects (comma-separated)."""
    subjects_added = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        teacher_name = str(row[0]).strip()
        subjects_str = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if not teacher_name or not subjects_str:
            continue

        subject_list = [s.strip() for s in subjects_str.split(",") if s.strip()]

        for s in subject_list:
            if s not in subjects_added:
                existing = Subject.query.filter_by(name=s, user_id=current_user.id).first()
                if not existing:
                    db.session.add(Subject(name=s, user_id=current_user.id))
                subjects_added.add(s)

        existing_t = Teacher.query.filter_by(name=teacher_name, user_id=current_user.id).first()
        if not existing_t:
            teacher = Teacher(name=teacher_name, user_id=current_user.id)
            db.session.add(teacher)
            db.session.flush()
            for s in subject_list:
                db.session.add(TeacherSubject(teacher_id=teacher.id, subject_name=s))


def _process_class_section_sheet(sheet):
    """Process Excel sheet: Column A = Class Name, Column B = Sections (comma-separated)."""
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        class_name = str(row[0]).strip()
        sections_str = str(row[1]).strip() if len(row) > 1 and row[1] else "A"
        if not class_name:
            continue

        sections = [s.strip() for s in sections_str.split(",") if s.strip()]
        for section in sections:
            existing = ClassSection.query.filter_by(
                class_name=class_name, section=section, user_id=current_user.id
            ).first()
            if not existing:
                db.session.add(ClassSection(
                    class_name=class_name, section=section, user_id=current_user.id
                ))


@app.route("/generate")
@login_required
def generate():
    """Generate the routine and display results."""
    subjects = Subject.query.filter_by(user_id=current_user.id).all()
    teachers = Teacher.query.filter_by(user_id=current_user.id).all()
    classes = ClassSection.query.filter_by(user_id=current_user.id).all()
    settings = get_or_create_settings(current_user.id)

    if not subjects:
        flash("No subjects added. Please add subjects first.", "danger")
        return redirect(url_for("index"))
    if not teachers:
        flash("No teachers added. Please add teachers first.", "danger")
        return redirect(url_for("index"))
    if not classes:
        flash("No classes added. Please add classes first.", "danger")
        return redirect(url_for("index"))

    # Build generator input
    # Load class-subject period limits
    csp_records = ClassSubjectPeriod.query.filter_by(user_id=current_user.id).all()
    class_subject_periods = {}
    for rec in csp_records:
        class_subject_periods.setdefault(rec.class_name, {})[rec.subject_name] = rec.periods_per_week

    # Class teacher assignments
    class_teachers_map = {}
    if settings.enable_class_teacher:
        ct_records = ClassTeacher.query.filter_by(user_id=current_user.id).all()
        for rec in ct_records:
            class_teachers_map[rec.class_name] = {
                "teacher": rec.teacher_name,
                "subject": rec.subject_name,
            }

    input_data = SchoolInputData(
        subjects=[s.name for s in subjects],
        teachers=[TeacherData(name=t.name, subjects=t.subject_names) for t in teachers],
        class_keys=[c.full_name for c in classes],
        periods_per_day=settings.periods_per_day,
        days=settings.days_list,
        day_periods=settings.day_periods,
        first_period_subjects=settings.first_period_subjects,
        class_subject_periods=class_subject_periods,
        class_teachers=class_teachers_map,
        max_teacher_periods_per_day=settings.max_teacher_periods_per_day or 0,
        avoid_consecutive=settings.avoid_consecutive,
    )

    gen = RoutineGenerator(input_data)
    success = gen.generate(max_attempts=500)

    if not success:
        flash("Could not generate a conflict-free routine. Try adding more teachers or reducing classes/periods.", "danger")
        return redirect(url_for("index"))

    routine_dict = gen.get_routine_dict()

    # Save to database
    GeneratedRoutine.query.filter_by(user_id=current_user.id).delete()
    gen_routine = GeneratedRoutine(user_id=current_user.id)
    gen_routine.routine_data = routine_dict
    db.session.add(gen_routine)
    db.session.commit()

    flash("Routine generated successfully!", "success")
    return redirect(url_for("index", tab="routine"))


@app.route("/save_routine", methods=["POST"])
@login_required
def save_routine():
    """Validate and save an edited routine via AJAX. Returns JSON response."""
    from flask import jsonify

    data = request.get_json()
    if not data or "routine" not in data:
        return jsonify({"success": False, "error": "No routine data provided."}), 400

    routine = data["routine"]
    settings = get_or_create_settings(current_user.id)
    teachers = Teacher.query.filter_by(user_id=current_user.id).all()
    subjects = Subject.query.filter_by(user_id=current_user.id).all()

    teacher_names = {t.name for t in teachers}
    subject_names = {s.name for s in subjects}
    # Build teacher->subjects map for validation
    teacher_subject_map = {t.name: set(t.subject_names) for t in teachers}

    errors = []

    # Validate each day/period slot
    for class_name, schedule in routine.items():
        for day, periods in schedule.items():
            num_expected = settings.get_periods_for_day(day)
            if len(periods) != num_expected:
                errors.append(f"{class_name} / {day}: Expected {num_expected} periods, got {len(periods)}.")

            for idx, period in enumerate(periods):
                if period is None:
                    errors.append(f"{class_name} / {day} / Period {idx+1}: Empty slot not allowed.")
                    continue
                subj = period.get("subject", "")
                tchr = period.get("teacher", "")
                if not subj or not tchr:
                    errors.append(f"{class_name} / {day} / Period {idx+1}: Subject and teacher are required.")
                    continue
                if subj not in subject_names:
                    errors.append(f"{class_name} / {day} / Period {idx+1}: Unknown subject '{subj}'.")
                if tchr not in teacher_names:
                    errors.append(f"{class_name} / {day} / Period {idx+1}: Unknown teacher '{tchr}'.")
                elif subj in subject_names and tchr in teacher_names:
                    if subj not in teacher_subject_map.get(tchr, set()):
                        errors.append(f"{class_name} / {day} / Period {idx+1}: '{tchr}' cannot teach '{subj}'.")

    # Check teacher conflicts (same teacher in two classes at same time)
    for day in settings.days_list:
        num_periods = settings.get_periods_for_day(day)
        for idx in range(num_periods):
            teacher_slot = {}  # teacher_name -> class_name
            for class_name, schedule in routine.items():
                periods = schedule.get(day, [])
                if idx < len(periods) and periods[idx]:
                    tchr = periods[idx].get("teacher", "")
                    if tchr in teacher_slot:
                        errors.append(
                            f"{day} / Period {idx+1}: Teacher '{tchr}' assigned to both "
                            f"'{teacher_slot[tchr]}' and '{class_name}'."
                        )
                    else:
                        teacher_slot[tchr] = class_name

    if errors:
        return jsonify({"success": False, "errors": errors}), 400

    # All valid — save
    GeneratedRoutine.query.filter_by(user_id=current_user.id).delete()
    gen_routine = GeneratedRoutine(user_id=current_user.id)
    gen_routine.routine_data = routine
    db.session.add(gen_routine)
    db.session.commit()

    return jsonify({"success": True, "message": "Routine saved successfully!"})


@app.route("/substitute", methods=["POST"])
@login_required
def substitute():
    """Find substitute teachers and redirect back to home page substitute tab."""
    teachers = Teacher.query.filter_by(user_id=current_user.id).all()
    settings = get_or_create_settings(current_user.id)
    gen_routine = GeneratedRoutine.query.filter_by(user_id=current_user.id).first()

    if not gen_routine:
        flash("Please generate a routine first before checking substitutes.", "warning")
        return redirect(url_for("index"))

    routine = gen_routine.routine_data
    selected_teacher = request.form.get("absent_teacher", "").strip()
    selected_day = request.form.get("day", "").strip()

    if not selected_teacher or not selected_day:
        flash("Please select a teacher and day.", "warning")
        return redirect(url_for("index", tab="substitute"))

    absent_teacher_obj = next((t for t in teachers if t.name == selected_teacher), None)
    if not absent_teacher_obj:
        flash("Teacher not found.", "danger")
        return redirect(url_for("index", tab="substitute"))

    absent_subjects = absent_teacher_obj.subject_names
    num_periods = settings.get_periods_for_day(selected_day)

    result = []
    for period_idx in range(num_periods):
        busy_teachers = set()
        absent_teacher_classes = []

        for class_name, schedule in routine.items():
            day_schedule = schedule.get(selected_day, [])
            if period_idx < len(day_schedule) and day_schedule[period_idx]:
                period_data = day_schedule[period_idx]
                busy_teachers.add(period_data["teacher"])
                if period_data["teacher"] == selected_teacher:
                    absent_teacher_classes.append({
                        "class": class_name,
                        "subject": period_data["subject"],
                    })

        available_subs = []
        for teacher in teachers:
            if teacher.name == selected_teacher:
                continue
            if teacher.name in busy_teachers:
                continue
            common_subjects = [s for s in teacher.subject_names if s in absent_subjects]
            if common_subjects:
                available_subs.append({
                    "name": teacher.name,
                    "can_teach": common_subjects,
                })

        free_others = []
        for teacher in teachers:
            if teacher.name == selected_teacher:
                continue
            if teacher.name in busy_teachers:
                continue
            common_subjects = [s for s in teacher.subject_names if s in absent_subjects]
            if not common_subjects:
                free_others.append({
                    "name": teacher.name,
                    "can_teach": teacher.subject_names,
                })

        result.append({
            "period": period_idx + 1,
            "absent_classes": absent_teacher_classes,
            "substitutes": available_subs,
            "free_others": free_others,
        })

    # Store in session for display
    session["substitute_result"] = result
    session["substitute_teacher"] = selected_teacher
    session["substitute_day"] = selected_day

    return redirect(url_for("index", tab="substitute"))


@app.route("/clear")
@login_required
def clear_data():
    """Clear all data for current user."""
    Subject.query.filter_by(user_id=current_user.id).delete()
    Teacher.query.filter_by(user_id=current_user.id).delete()
    ClassSection.query.filter_by(user_id=current_user.id).delete()
    GeneratedRoutine.query.filter_by(user_id=current_user.id).delete()
    ClassSubjectPeriod.query.filter_by(user_id=current_user.id).delete()
    ClassTeacher.query.filter_by(user_id=current_user.id).delete()
    settings = ScheduleSettings.query.filter_by(user_id=current_user.id).first()
    if settings:
        db.session.delete(settings)
    db.session.commit()
    flash("All data has been deleted successfully.", "success")
    return redirect(url_for("index"))


# ─── DB Init ─────────────────────────────────────────────────

with app.app_context():
    db.create_all()


if __name__ == "__main__":
    app.run(debug=True, port=5000)
